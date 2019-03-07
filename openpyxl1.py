import openpyxl
import os
from openpyxl import Workbook

wb=openpyxl.load_workbook("D:\\test.xlsx")
#print(wb)
#print(wb.get_sheet_names)

ws1=wb.active
#ws2=wb.create_sheet("my new sheet")

#print(ws1.title)
#print(ws2.title)

#clone=wb.copy_worksheet(ws1)
#clone.title="cloned sheet"

print("all sheets available are ",wb.sheetnames)

print("max columns are ",ws1.max_column," max rows are ",ws1.max_row)

roll=ws1["B1"]
#print("roll number is ",roll.value)

msg=ws1.cell(row=1,column=1)
#print(msg.value)


for each_row in range(1,ws1.max_row+1):
    for each_column in range(1,ws1.max_column+1):
        value1=ws1.cell(row=each_row,column=each_column)
        print(value1.value+" ",end=" ")
    print("\n")


#for item in ws1.iter_rows(min_row=1,max_col=3,max_row=2,values_only=True):
#    for every in item:
#       print(item)


#for row in ws1.values:
#    for each in row:
#        print(each)

roll2=ws1["B2"]
roll2=ws1.cell(row=2,column=2)
roll2.value="y16cs964"

roll=str(input("enter roll number to search"))

for each_row in range(1,ws1.max_row+1):
    for each_column in range(1,ws1.max_column+1):
        rollnumber=ws1.cell(row=each_row,column=each_column)
        if rollnumber.value==roll:
            print("Found")
            for every in range(1,ws1.max_column+1):
                print(ws1.cell(row=each_row,column=every).value+" ",end=" ")


wb.save("D:\\test.xlsx")

#print(ws)
